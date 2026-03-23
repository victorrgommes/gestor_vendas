import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from decimal import Decimal

class VendasExporter:
    """Classe para gerenciar vendas e exportação para Excel"""
    
    def __init__(self):
        self.vendas = []
        self.colunas = ["ID", "Produto", "Quantidade", "Valor Unitário", "Valor Total", "Data"]
    
    def adicionar_venda(self, produto, quantidade, valor_unitario, data=None):
        """Adiciona uma venda à lista"""
        if data is None:
            data = datetime.now().strftime("%d/%m/%Y")
        
        venda = {
            "id": len(self.vendas) + 1,
            "produto": produto,
            "quantidade": quantidade,
            "valor_unitario": valor_unitario,
            "valor_total": quantidade * valor_unitario,
            "data": data
        }
        self.vendas.append(venda)
        return venda
    
    def remover_venda(self, id_venda):
        """Remove uma venda pelo ID"""
        self.vendas = [v for v in self.vendas if v["id"] != id_venda]
        # Reorganizar IDs
        for idx, venda in enumerate(self.vendas, 1):
            venda["id"] = idx
    
    def exportar_excel(self, nome_arquivo="relatorio_vendas.xlsx"):
        """Exporta as vendas para um arquivo Excel"""
        if not self.vendas:
            return False
        
        try:
            # Criar novo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vendas"
            
            # Adicionar cabeçalho
            for col, coluna in enumerate(self.colunas, 1):
                celula = ws.cell(row=1, column=col, value=coluna)
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                celula.alignment = Alignment(horizontal="center", vertical="center")
            
            # Adicionar dados
            for row, venda in enumerate(self.vendas, 2):
                ws.cell(row=row, column=1, value=venda["id"])
                ws.cell(row=row, column=2, value=venda["produto"])
                ws.cell(row=row, column=3, value=venda["quantidade"])
                ws.cell(row=row, column=4, value=f"R$ {venda['valor_unitario']:.2f}")
                ws.cell(row=row, column=5, value=f"R$ {venda['valor_total']:.2f}")
                ws.cell(row=row, column=6, value=venda["data"])
                
                # Alinhar dados
                for col in range(1, 7):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            
            # Adicionar linha de total
            total_row = len(self.vendas) + 2
            ws.cell(row=total_row, column=2, value="TOTAL:")
            ws.cell(row=total_row, column=2).font = Font(bold=True)
            
            total_valor = sum(v["valor_total"] for v in self.vendas)
            total_cell = ws.cell(row=total_row, column=5, value=f"R$ {total_valor:.2f}")
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            
            # Salvar arquivo
            caminho_arquivo = Path(__file__).parent / nome_arquivo
            wb.save(caminho_arquivo)
            return str(caminho_arquivo)
        except Exception as e:
            return None


class InterfaceVendas:
    """Interface gráfica para gerenciar vendas"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Vendas - Exportar para Excel")
        self.root.geometry("900x650")
        self.root.configure(bg="#f0f0f0")
        
        self.exporter = VendasExporter()
        self.criar_interface()
    
    def criar_interface(self):
        """Cria todos os elementos da interface"""
        
        # ===== PAINEL SUPERIOR (Entrada de dados) =====
        frame_entrada = ttk.LabelFrame(self.root, text="Adicionar Nova Venda", padding=10)
        frame_entrada.pack(fill=tk.X, padx=10, pady=10)
        
        # Produto
        ttk.Label(frame_entrada, text="Produto:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_produto = ttk.Entry(frame_entrada, width=30)
        self.entry_produto.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=5)
        
        # Quantidade
        ttk.Label(frame_entrada, text="Quantidade:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        self.entry_quantidade = ttk.Entry(frame_entrada, width=12)
        self.entry_quantidade.grid(row=0, column=3, sticky=tk.EW, padx=5, pady=5)
        
        # Valor Unitário
        ttk.Label(frame_entrada, text="Valor Unit. (R$):").grid(row=0, column=4, sticky=tk.W, padx=5, pady=5)
        self.entry_valor = ttk.Entry(frame_entrada, width=15)
        self.entry_valor.grid(row=0, column=5, sticky=tk.EW, padx=5, pady=5)
        
        # Botão Adicionar
        btn_adicionar = ttk.Button(frame_entrada, text="Adicionar", command=self.adicionar_venda)
        btn_adicionar.grid(row=0, column=6, sticky=tk.EW, padx=5, pady=5)
        
        frame_entrada.columnconfigure(1, weight=1)
        frame_entrada.columnconfigure(3, weight=1)
        frame_entrada.columnconfigure(5, weight=1)
        
        # ===== PAINEL CENTRAL (Tabela de vendas) =====
        frame_tabela = ttk.LabelFrame(self.root, text="Vendas Registradas", padding=10)
        frame_tabela.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Criar Treeview
        self.tree = ttk.Treeview(frame_tabela, columns=("ID", "Produto", "Qtd", "Valor Unit.", "Total"), 
                                  height=15, show="headings")
        
        # Definir colunas
        self.tree.column("ID", width=40, anchor=tk.CENTER)
        self.tree.column("Produto", width=300, anchor=tk.W)
        self.tree.column("Qtd", width=80, anchor=tk.CENTER)
        self.tree.column("Valor Unit.", width=120, anchor=tk.CENTER)
        self.tree.column("Total", width=120, anchor=tk.CENTER)
        
        self.tree.heading("ID", text="ID")
        self.tree.heading("Produto", text="Produto")
        self.tree.heading("Qtd", text="Quantidade")
        self.tree.heading("Valor Unit.", text="Valor Unitário")
        self.tree.heading("Total", text="Total")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_tabela, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # ===== PAINEL INFERIOR (Resumo e Botões) =====
        frame_resumo = ttk.Frame(self.root)
        frame_resumo.pack(fill=tk.X, padx=10, pady=5)
        
        # Label de total
        self.label_total = ttk.Label(frame_resumo, text="Total: R$ 0,00", font=("Arial", 12, "bold"))
        self.label_total.pack(side=tk.LEFT, padx=10)
        
        # Botões de ação
        frame_botoes = ttk.Frame(self.root)
        frame_botoes.pack(fill=tk.X, padx=10, pady=10)
        
        btn_exportar = ttk.Button(frame_botoes, text="📊 Exportar para Excel", command=self.exportar_excel)
        btn_exportar.pack(side=tk.LEFT, padx=5)
        
        btn_remover = ttk.Button(frame_botoes, text="🗑️ Remover Selecionado", command=self.remover_venda)
        btn_remover.pack(side=tk.LEFT, padx=5)
        
        btn_limpar = ttk.Button(frame_botoes, text="🔄 Limpar Tudo", command=self.limpar_todas)
        btn_limpar.pack(side=tk.LEFT, padx=5)
        
        btn_sair = ttk.Button(frame_botoes, text="❌ Sair", command=self.root.quit)
        btn_sair.pack(side=tk.RIGHT, padx=5)
    
    def adicionar_venda(self):
        """Adiciona uma venda da interface"""
        try:
            produto = self.entry_produto.get().strip()
            quantidade = int(self.entry_quantidade.get())
            valor_unitario = float(self.entry_valor.get().replace(",", "."))
            
            if not produto:
                messagebox.showwarning("Aviso", "Por favor, insira o nome do produto!")
                return
            
            if quantidade <= 0:
                messagebox.showwarning("Aviso", "Quantidade deve ser maior que 0!")
                return
            
            if valor_unitario <= 0:
                messagebox.showwarning("Aviso", "Valor deve ser maior que 0!")
                return
            
            # Adicionar à lista
            venda = self.exporter.adicionar_venda(produto, quantidade, valor_unitario)
            
            # Adicionar à tabela
            total = venda["valor_total"]
            self.tree.insert("", tk.END, values=(
                venda["id"],
                produto,
                quantidade,
                f"R$ {valor_unitario:.2f}",
                f"R$ {total:.2f}"
            ))
            
            # Limpar campos
            self.entry_produto.delete(0, tk.END)
            self.entry_quantidade.delete(0, tk.END)
            self.entry_valor.delete(0, tk.END)
            self.entry_produto.focus()
            
            # Atualizar total
            self.atualizar_total()
            messagebox.showinfo("Sucesso", f"Venda adicionada: {produto}")
            
        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira valores válidos!\nQuantidade e Valor devem ser números.")
    
    def remover_venda(self):
        """Remove a venda selecionada"""
        selecionado = self.tree.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione uma venda para remover!")
            return
        
        item = selecionado[0]
        valores = self.tree.item(item, "values")
        id_venda = int(valores[0])
        
        if messagebox.askyesno("Confirmar", f"Remover venda ID {id_venda}?"):
            self.tree.delete(item)
            self.exporter.remover_venda(id_venda)
            self.atualizar_total()
            messagebox.showinfo("Sucesso", "Venda removida!")
    
    def limpar_todas(self):
        """Limpa todas as vendas"""
        if not self.exporter.vendas:
            messagebox.showinfo("Aviso", "Não há vendas para limpar!")
            return
        
        if messagebox.askyesno("Confirmar", "Deseja limpar TODAS as vendas?"):
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.exporter.vendas = []
            self.atualizar_total()
            messagebox.showinfo("Sucesso", "Todas as vendas foram removidas!")
    
    def atualizar_total(self):
        """Atualiza o total exibido"""
        total = sum(v["valor_total"] for v in self.exporter.vendas)
        self.label_total.config(text=f"Total: R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    def exportar_excel(self):
        """Exporta as vendas para Excel"""
        if not self.exporter.vendas:
            messagebox.showwarning("Aviso", "Nenhuma venda para exportar!")
            return
        
        # Caixa de diálogo para salvar
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"relatorio_vendas_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        )
        
        if not arquivo:
            return
        
        resultado = self.exporter.exportar_excel(arquivo)
        
        if resultado:
            messagebox.showinfo("Sucesso", f"Arquivo exportado!\n{resultado}")
        else:
            messagebox.showerror("Erro", "Erro ao exportar arquivo!")


def main():
    """Função principal"""
    root = tk.Tk()
    app = InterfaceVendas(root)
    root.mainloop()


if __name__ == "__main__":
    main()
